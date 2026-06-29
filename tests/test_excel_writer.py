from pathlib import Path

import openpyxl


def test_excel_writer_outputs_fixed_columns_logs_and_evidence(tmp_path: Path):
    from columns import TARGET_COLUMNS
    from excel_writer import EVIDENCE_SHEET, LOG_SHEET, RESULT_SHEET, write_excel

    output = tmp_path / "result.xlsx"
    records = [{"info_id": "A-1", "文章时间": "2026-05-22", "报销比例": "70%"}]
    logs = [
        {
            "source_id": "A-1",
            "info_id": "A-1",
            "title": "医保政策",
            "status": "success",
            "message": "",
            "records": 1,
            "llm_used": True,
            "llm_success": True,
            "need_manual_review": False,
            "review_reason": "",
        }
    ]
    evidence = [
        {
            "source_id": "A-1",
            "info_id": "A-1",
            "field": "报销比例",
            "value": "70%",
            "evidence": "报销比例70%",
            "confidence": 0.95,
            "source": "llm",
            "rule_name": "deepseek-v4-flash",
        }
    ]

    write_excel(records, logs, output, TARGET_COLUMNS, field_evidence=evidence)

    wb = openpyxl.load_workbook(output)
    assert RESULT_SHEET in wb.sheetnames
    assert LOG_SHEET in wb.sheetnames
    assert EVIDENCE_SHEET in wb.sheetnames
    assert [cell.value for cell in wb[RESULT_SHEET][1]] == TARGET_COLUMNS
    assert [cell.value for cell in wb[LOG_SHEET][1]][:15] == [
        "source_id",
        "info_id",
        "title",
        "status",
        "message",
        "records",
        "time",
        "llm_used",
        "llm_success",
        "llm_error",
        "need_manual_review",
        "review_reason",
        "table_count",
        "table_row_count",
        "extraction_mode",
    ]
    assert [cell.value for cell in wb[EVIDENCE_SHEET][1]] == [
        "source_id",
        "info_id",
        "row_index",
        "field",
        "value",
        "evidence",
        "confidence",
        "source",
        "rule_name",
    ]
    assert wb[EVIDENCE_SHEET]["D2"].value == "报销比例"
