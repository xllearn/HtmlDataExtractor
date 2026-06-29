import json
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
import yaml


def write_yaml(path: Path, data: dict) -> Path:
    path.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8")
    return path


def field_config(path: Path) -> Path:
    from columns import TARGET_COLUMNS

    return write_yaml(
        path,
        {
            "output_columns": TARGET_COLUMNS,
            "default_missing_value": "",
            "field_missing_values": {
                "个人账户计入办法": "-",
                "个人账户使用范围": "-",
                "病种名称": "-",
                "类型": "-",
                "标化类型": "-",
                "保险类型": "-",
                "人员类型": "-",
                "病种类型": "-",
                "就诊地域": "-",
                "医院类型": "-",
                "就诊情况": "-",
                "区间": "-",
                "起付标准": "-",
                "补助限额": "-",
                "报销比例": "-",
                "审核状态0待审核1已审核": "0",
                "是否需要手动修改执行状态(1是0否)": "0",
            },
        },
    )


def test_db_reader_loads_dynamic_columns_from_sqlite(tmp_path: Path):
    from sqlalchemy import create_engine, text

    from db_reader import load_db_config, read_records

    db_path = tmp_path / "articles.sqlite"
    engine = create_engine(f"sqlite:///{db_path}")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE policy_article (
                    id INTEGER PRIMARY KEY,
                    article_id TEXT,
                    title TEXT,
                    html_content TEXT,
                    content TEXT,
                    publish_time TEXT,
                    audit_time TEXT,
                    region TEXT,
                    related_info TEXT,
                    unused_column TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO policy_article
                (id, article_id, title, html_content, content, publish_time, audit_time, region, related_info, unused_column)
                VALUES
                (1, 'A-1', '西安医保政策', '<p>医保 报销比例：70%</p>', 'fallback', '2026-05-22', '2026-05-23', '西安市', '来源A', 'ignore')
                """
            )
        )

    config_path = write_yaml(
        tmp_path / "db.yml",
        {
            "database": {"url": f"sqlite:///{db_path}"},
            "source": {
                "table": "policy_article",
                "id_column": "id",
                "info_id_column": "article_id",
                "title_column": "title",
                "html_column": "html_content",
                "text_column": "content",
                "article_time_column": "publish_time",
                "audit_time_column": "audit_time",
                "region_column": "region",
                "related_info_column": "related_info",
            },
            "query": {"where": "id = 1", "limit": 1, "order_by": "id ASC"},
        },
    )

    config = load_db_config(config_path)
    records = read_records(config)

    assert len(records) == 1
    record = records[0]
    assert record["source_id"] == "1"
    assert record["info_id"] == "A-1"
    assert record["title"] == "西安医保政策"
    assert record["html"] == "<p>医保 报销比例：70%</p>"
    assert record["text"] == "fallback"
    assert record["article_time"] == "2026-05-22"
    assert record["audit_time"] == "2026-05-23"
    assert record["region"] == "西安市"
    assert record["related_info"] == "来源A"
    assert "unused_column" not in record["raw"]


def test_extractor_uses_db_fields_first_and_skips_keyword_misses(tmp_path: Path):
    from columns import TARGET_COLUMNS
    from extractor import extract_record

    record = {
        "source_id": "1",
        "info_id": "DB-INFO",
        "title": "西安医保待遇",
        "html": """
            <table>
              <tr><th>类型</th><th>报销比例</th><th>起付标准</th></tr>
              <tr><td>门诊统筹</td><td>70%</td><td>200元</td></tr>
            </table>
        """,
        "text": "医保 门诊 报销比例：80% 审核日期：2026年5月24日",
        "article_time": "2026-05-22 10:00:00",
        "audit_time": "2026-05-23",
        "region": "数据库地区",
        "related_info": "数据库来源",
        "raw": {},
    }

    records, log = extract_record(record, TARGET_COLUMNS, keyword="医保")

    assert log["status"] == "success"
    assert len(records) == 1
    result = records[0]
    assert list(result.keys()) == TARGET_COLUMNS
    assert result["info_id"] == "DB-INFO"
    assert result["文章时间"] == "2026-05-22"
    assert result["审核日期"] == "2026-05-23"
    assert result["地区名称"] == "数据库地区"
    assert result["相关资讯"] == "数据库来源"
    assert result["类型"] == "门诊统筹"
    assert result["报销比例"] == "70%"
    assert result["起付标准"] == "200元"

    skipped, skipped_log = extract_record(record, TARGET_COLUMNS, keyword="完全不匹配")
    assert skipped == []
    assert skipped_log["status"] == "skipped_keyword"


def test_cleaner_fills_deduplicates_and_sorts(tmp_path: Path):
    from cleaner import clean_records, load_field_mapping

    mapping = load_field_mapping(field_config(tmp_path / "fields.yml"))
    rows = [
        {"文章时间": "2026年5月22日", "info_id": "2", "地区名称": "西安市", "类型": "门诊", "报销比例": "70%"},
        {"文章时间": "2026-01-01", "info_id": "1", "地区名称": "西安市", "类型": "住院", "报销比例": "80%"},
        {"文章时间": "2026年5月22日", "info_id": "2", "地区名称": "西安市", "类型": "门诊", "报销比例": "70%"},
    ]

    cleaned = clean_records(rows, mapping)

    assert len(cleaned) == 2
    assert cleaned[0]["info_id"] == "1"
    assert cleaned[0]["病种名称"] == "-"
    assert cleaned[0]["审核状态0待审核1已审核"] == "0"
    assert cleaned[0]["是否需要手动修改执行状态(1是0否)"] == "0"
    assert cleaned[1]["文章时间"] == "2026-05-22"


def test_excel_writer_outputs_only_fixed_columns_and_logs(tmp_path: Path):
    from columns import TARGET_COLUMNS
    from excel_writer import write_excel

    output = tmp_path / "result.xlsx"
    logs = [{"source_id": "1", "info_id": "A-1", "title": "标题", "status": "success", "message": "", "records": 1}]
    write_excel(
        [{"info_id": "A-1", "文章时间": "2026-05-22", "地区名称": "西安市", "类型": "门诊"}],
        logs,
        output,
        TARGET_COLUMNS,
    )

    wb = openpyxl.load_workbook(output)
    assert "结果数据" in wb.sheetnames
    assert "采集日志" in wb.sheetnames
    result_headers = [cell.value for cell in wb["结果数据"][1]]
    assert result_headers == TARGET_COLUMNS
    log_headers = [cell.value for cell in wb["采集日志"][1]]
    assert log_headers[:4] == ["source_id", "info_id", "title", "status"]


def test_run_pipeline_and_api_sample(tmp_path: Path):
    from sqlalchemy import create_engine, text
    from fastapi.testclient import TestClient

    from app import app
    from columns import TARGET_COLUMNS
    from main import run_pipeline

    db_path = tmp_path / "pipeline.sqlite"
    engine = create_engine(f"sqlite:///{db_path}")
    with engine.begin() as conn:
        conn.execute(text("CREATE TABLE policy_article (id INTEGER PRIMARY KEY, title TEXT, content TEXT, publish_time TEXT, region TEXT)"))
        conn.execute(
            text(
                "INSERT INTO policy_article (id, title, content, publish_time, region) "
                "VALUES (1, '示例政策', '医保 类型：门诊 报销比例：70%', '2026-05-22', '西安市')"
            )
        )

    db_config = write_yaml(
        tmp_path / "db.yml",
        {
            "database": {"url": f"sqlite:///{db_path}"},
            "source": {
                "table": "policy_article",
                "id_column": "id",
                "info_id_column": "",
                "title_column": "title",
                "html_column": "",
                "text_column": "content",
                "article_time_column": "publish_time",
                "audit_time_column": "",
                "region_column": "region",
                "related_info_column": "",
            },
            "query": {"where": "", "limit": 10, "order_by": "id ASC"},
        },
    )
    output = tmp_path / "out.xlsx"
    summary = run_pipeline(db_config, field_config(tmp_path / "fields.yml"), output, keyword="医保", template_path="")

    assert summary["read_count"] == 1
    assert summary["success_count"] == 1
    assert summary["failed_count"] == 0
    assert Path(summary["output_file"]).exists()
    df = pd.read_excel(output, sheet_name="结果数据")
    assert list(df.columns) == TARGET_COLUMNS
    assert len(df) == 1

    client = TestClient(app)
    response = client.get("/api/sample")
    assert response.status_code == 200
    sample = response.json()
    assert sample["record"]["source_id"]
    assert list(sample["result"].keys()) == TARGET_COLUMNS
