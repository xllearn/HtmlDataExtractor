from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, Sequence

from openpyxl import Workbook, load_workbook

from columns import TARGET_COLUMNS

RESULT_SHEET = "结果数据"
LOG_SHEET = "采集日志"
EVIDENCE_SHEET = "字段证据"
LOG_COLUMNS = [
    "source_id",
    "info_id",
    "title",
    "status",
    "message",
    "records",
    "time",
    "llm_used",
    "llm_success",
    "need_manual_review",
    "review_reason",
]
EVIDENCE_COLUMNS = ["source_id", "info_id", "field", "value", "evidence", "confidence", "source", "rule_name"]


def ensure_sheet(workbook, name: str):
    if name in workbook.sheetnames:
        return workbook[name]
    return workbook.create_sheet(name)


def clear_sheet(sheet) -> None:
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)


def clear_data_rows(sheet) -> None:
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)


def copy_row_style(sheet, source_row: int, target_row: int, column_count: int) -> None:
    for col in range(1, column_count + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.border:
            target.border = copy(source.border)
        if source.fill:
            target.fill = copy(source.fill)
        if source.font:
            target.font = copy(source.font)


def snapshot_row_style(sheet, source_row: int, column_count: int) -> list[Dict[str, Any]]:
    styles = []
    for col in range(1, column_count + 1):
        source = sheet.cell(source_row, col)
        styles.append(
            {
                "style": copy(source._style) if source.has_style else None,
                "number_format": source.number_format,
                "alignment": copy(source.alignment) if source.alignment else None,
                "border": copy(source.border) if source.border else None,
                "fill": copy(source.fill) if source.fill else None,
                "font": copy(source.font) if source.font else None,
            }
        )
    return styles


def apply_row_style(target, style: Dict[str, Any]) -> None:
    if style["style"] is not None:
        target._style = copy(style["style"])
    if style["number_format"]:
        target.number_format = style["number_format"]
    if style["alignment"]:
        target.alignment = copy(style["alignment"])
    if style["border"]:
        target.border = copy(style["border"])
    if style["fill"]:
        target.fill = copy(style["fill"])
    if style["font"]:
        target.font = copy(style["font"])


def apply_cached_row_style(sheet, styles: Sequence[Dict[str, Any]], target_row: int) -> None:
    for col, style in enumerate(styles, start=1):
        apply_row_style(sheet.cell(target_row, col), style)


def write_plain_rows(sheet, headers: Sequence[str], rows: Iterable[Dict[str, Any]]) -> None:
    clear_sheet(sheet)
    sheet.append(list(headers))
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def write_template_result_rows(sheet, headers: Sequence[str], rows: Iterable[Dict[str, Any]]) -> None:
    if sheet.max_row < 1:
        sheet.append(list(headers))
    else:
        for index, header in enumerate(headers, start=1):
            sheet.cell(1, index).value = header
    style_source_row = 2 if sheet.max_row >= 2 else 1
    data_row_styles = snapshot_row_style(sheet, style_source_row, len(headers))
    clear_data_rows(sheet)
    for row_index, row in enumerate(rows, start=2):
        apply_cached_row_style(sheet, data_row_styles, row_index)
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row_index, col_index).value = row.get(header, "")


def write_excel(
    records: Iterable[Dict[str, Any]],
    logs: Iterable[Dict[str, Any]],
    output: str | Path,
    columns: Sequence[str] | None = None,
    template_path: str | Path = "",
    field_evidence: Iterable[Dict[str, Any]] | None = None,
) -> str:
    columns = list(columns or TARGET_COLUMNS)
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    template = Path(template_path) if template_path else None
    use_template = bool(template and template.exists())
    if use_template:
        workbook = load_workbook(template)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)

    result_sheet = ensure_sheet(workbook, RESULT_SHEET)
    log_sheet = ensure_sheet(workbook, LOG_SHEET)
    evidence_sheet = ensure_sheet(workbook, EVIDENCE_SHEET)
    if use_template:
        write_template_result_rows(result_sheet, columns, records)
    else:
        write_plain_rows(result_sheet, columns, records)

    prepared_logs = []
    prepared_evidence = [dict(row) for row in (field_evidence or [])]
    for log in logs:
        item = dict(log)
        item.setdefault("time", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        item.setdefault("llm_used", False)
        item.setdefault("llm_success", False)
        item.setdefault("need_manual_review", False)
        item.setdefault("review_reason", "")
        prepared_logs.append(item)
        prepared_evidence.extend(dict(row) for row in item.get("field_evidence") or [])
    write_plain_rows(log_sheet, LOG_COLUMNS, prepared_logs)
    write_plain_rows(evidence_sheet, EVIDENCE_COLUMNS, prepared_evidence)
    workbook.save(output_path)
    return str(output_path)
